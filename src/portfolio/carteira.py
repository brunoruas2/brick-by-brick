"""
Gestao de posicoes e movimentacoes da carteira de FIIs.

Logica de preco medio:
  Compra: preco_medio = (cotas_antigas * pm_antigo + cotas_novas * preco) / total_cotas
  Venda: preco_medio permanece inalterado (apenas reduz cotas)
"""

import pandas as pd
from src.storage.database import connect


def add_compra(ticker: str, cotas: int, preco: float, data: str) -> None:
    """
    Registra uma compra e atualiza a posicao na carteira.
    Calcula preco medio ponderado se ja existir posicao ativa.
    """
    ticker = ticker.upper()
    valor_total = cotas * preco

    with connect() as conn:
        # Registra movimentacao
        conn.execute(
            """
            INSERT INTO movimentacoes (ticker, tipo, data, quantidade, preco_unitario, valor_total)
            VALUES (?, 'compra', ?, ?, ?, ?)
            """,
            (ticker, data, cotas, preco, valor_total),
        )

        # Busca CNPJ do ticker
        row = conn.execute(
            "SELECT cnpj FROM fiis WHERE ticker = ?", (ticker,)
        ).fetchone()
        cnpj = row["cnpj"] if row else None

        # Atualiza ou insere posicao na carteira
        existing = conn.execute(
            "SELECT id, cotas, preco_medio FROM carteira WHERE ticker = ? AND ativa = 1",
            (ticker,),
        ).fetchone()

        if existing:
            old_cotas = existing["cotas"]
            old_pm = existing["preco_medio"]
            new_cotas = old_cotas + cotas
            new_pm = (old_cotas * old_pm + cotas * preco) / new_cotas
            conn.execute(
                "UPDATE carteira SET cotas = ?, preco_medio = ? WHERE id = ?",
                (new_cotas, new_pm, existing["id"]),
            )
        else:
            conn.execute(
                """
                INSERT INTO carteira (ticker, cnpj, cotas, preco_medio, data_entrada, ativa)
                VALUES (?, ?, ?, ?, ?, 1)
                """,
                (ticker, cnpj, cotas, preco, data),
            )


def add_venda(ticker: str, cotas: int, preco: float, data: str) -> None:
    """
    Registra uma venda. Reduz cotas da posicao ativa.
    Se todas as cotas forem vendidas, marca ativa = 0.

    Raises:
        ValueError: se nao houver posicao ativa ou cotas insuficientes.
    """
    ticker = ticker.upper()
    valor_total = cotas * preco

    with connect() as conn:
        existing = conn.execute(
            "SELECT id, cotas FROM carteira WHERE ticker = ? AND ativa = 1",
            (ticker,),
        ).fetchone()

        if not existing:
            raise ValueError(f"Posicao ativa nao encontrada para {ticker}")
        if cotas > existing["cotas"]:
            raise ValueError(
                f"Cotas insuficientes: posicao tem {existing['cotas']}, venda requer {cotas}"
            )

        conn.execute(
            """
            INSERT INTO movimentacoes (ticker, tipo, data, quantidade, preco_unitario, valor_total)
            VALUES (?, 'venda', ?, ?, ?, ?)
            """,
            (ticker, data, cotas, preco, valor_total),
        )

        new_cotas = existing["cotas"] - cotas
        if new_cotas == 0:
            conn.execute(
                "UPDATE carteira SET cotas = 0, ativa = 0 WHERE id = ?",
                (existing["id"],),
            )
        else:
            conn.execute(
                "UPDATE carteira SET cotas = ? WHERE id = ?",
                (new_cotas, existing["id"]),
            )


def get_posicoes(month: str | None = None) -> pd.DataFrame:
    """
    Retorna DataFrame com posicoes ativas enriquecidas com dados de mercado.

    month: YYYY-MM. Se informado, usa precos e inf_mensal do mes indicado.
           Se None, usa os dados mais recentes disponiveis.

    Colunas:
      ticker, cotas, preco_medio, custo_total,
      nome, segmento,
      preco_atual, valor_atual,
      pl_capital, pl_pct,
      dy_mes, vpa, provento_est (cotas * vpa * dy_mes)
    """
    if month:
        cot_filter = "strftime('%Y-%m', data) <= ?"
        cot_params = (month,)
        im_filter  = "strftime('%Y-%m', data_referencia) <= ?"
        im_params  = (month,)
    else:
        cot_filter = "1=1"
        cot_params = ()
        im_filter  = "1=1"
        im_params  = ()

    with connect() as conn:
        rows = conn.execute(f"""
            SELECT
                c.ticker,
                c.cotas,
                c.preco_medio,
                ROUND(c.cotas * c.preco_medio, 2)          AS custo_total,
                f.nome,
                f.segmento,
                cot.fechamento                              AS preco_atual,
                ROUND(c.cotas * cot.fechamento, 2)         AS valor_atual,
                ROUND((cot.fechamento - c.preco_medio) * c.cotas, 2)          AS pl_capital,
                CASE WHEN c.preco_medio > 0
                     THEN ROUND((cot.fechamento - c.preco_medio) / c.preco_medio * 100, 2)
                     ELSE NULL END                         AS pl_pct,
                im.dy_mes,
                im.valor_patrimonial_cota                  AS vpa,
                ROUND(c.cotas * im.valor_patrimonial_cota * im.dy_mes, 2) AS provento_est
            FROM carteira c
            LEFT JOIN fiis f ON c.ticker = f.ticker
            LEFT JOIN (
                SELECT c1.ticker, c1.fechamento
                FROM cotacoes c1
                INNER JOIN (
                    SELECT ticker, MAX(data) AS max_data
                    FROM cotacoes
                    WHERE {cot_filter}
                    GROUP BY ticker
                ) c2 ON c1.ticker = c2.ticker AND c1.data = c2.max_data
            ) cot ON c.ticker = cot.ticker
            LEFT JOIN (
                SELECT f2.cnpj, im2.dy_mes, im2.valor_patrimonial_cota
                FROM inf_mensal im2
                INNER JOIN (
                    SELECT cnpj, MAX(data_referencia) AS max_ref
                    FROM inf_mensal
                    WHERE {im_filter}
                    GROUP BY cnpj
                ) latest ON im2.cnpj = latest.cnpj AND im2.data_referencia = latest.max_ref
                JOIN fiis f2 ON f2.cnpj = im2.cnpj
            ) im ON f.cnpj = im.cnpj
            WHERE c.ativa = 1
            ORDER BY c.ticker
        """, (*cot_params, *im_params)).fetchall()

    if not rows:
        return pd.DataFrame()

    return pd.DataFrame([dict(r) for r in rows])


def get_historico_dividendos(
    ticker: str | None = None,
    desde: str | None = None,
) -> pd.DataFrame:
    """
    Reconstroi historico de dividendos recebidos pela carteira, mes a mes.

    Logica:
      - Posicao mensal: reconstrói cotas e preco_medio a partir das movimentacoes
      - dividendo_cota  = dy_mes * preco_fechamento_do_mes
        (dy_mes ja esta em fracao decimal: 0.00661 = 0,66%)
        (fallback: VPA quando nao ha cotacao de mercado para aquele mes)
      - dividendo_recebido = cotas * dividendo_cota
      - yoc_mes = dividendo_recebido / custo_total * 100

    Colunas:
      mes, ticker, cotas, preco_medio, custo_total,
      preco_cota, dy_mes, dividendo_cota, dividendo_recebido, yoc_mes

    ticker: filtra por um unico ticker (opcional)
    desde:  YYYY-MM -- ignora meses anteriores a este (opcional)
    """
    ticker_upper = ticker.upper() if ticker else None

    with connect() as conn:
        # Posicoes ativas para obter o CNPJ de cada ticker
        if ticker_upper:
            pos_rows = conn.execute(
                "SELECT ticker, cnpj FROM carteira WHERE ticker = ? AND ativa = 1",
                (ticker_upper,),
            ).fetchall()
        else:
            pos_rows = conn.execute(
                "SELECT ticker, cnpj FROM carteira WHERE ativa = 1"
            ).fetchall()

        if not pos_rows:
            return pd.DataFrame()

        cnpj_por_ticker: dict[str, str | None] = {r["ticker"]: r["cnpj"] for r in pos_rows}
        tickers_list = list(cnpj_por_ticker.keys())
        ph = ",".join("?" * len(tickers_list))

        # Movimentacoes (compra/venda) para reconstruir posicao mensal
        movs_rows = conn.execute(
            f"""SELECT ticker, tipo, data, quantidade, preco_unitario
                FROM movimentacoes
                WHERE ticker IN ({ph}) AND tipo IN ('compra','venda')
                ORDER BY data, id""",
            tickers_list,
        ).fetchall()

        # Ultimo fechamento por ticker/mes
        cot_rows = conn.execute(
            f"""SELECT c.ticker, strftime('%Y-%m', c.data) AS mes, c.fechamento AS preco_cota
                FROM cotacoes c
                INNER JOIN (
                    SELECT ticker, strftime('%Y-%m', data) AS mes, MAX(data) AS max_data
                    FROM cotacoes WHERE ticker IN ({ph})
                    GROUP BY ticker, mes
                ) latest ON c.ticker = latest.ticker AND c.data = latest.max_data""",
            tickers_list,
        ).fetchall()

        # DY mensal e VPA por cnpj/mes
        cnpjs = [v for v in cnpj_por_ticker.values() if v]
        if cnpjs:
            ph_c = ",".join("?" * len(cnpjs))
            im_rows = conn.execute(
                f"""SELECT f.ticker,
                           strftime('%Y-%m', im.data_referencia) AS mes,
                           im.dy_mes,
                           im.valor_patrimonial_cota AS vpa
                    FROM inf_mensal im
                    JOIN fiis f ON f.cnpj = im.cnpj
                    WHERE im.cnpj IN ({ph_c}) AND im.dy_mes IS NOT NULL""",
                cnpjs,
            ).fetchall()
        else:
            im_rows = []

    if not movs_rows:
        return pd.DataFrame()

    movs = pd.DataFrame([dict(r) for r in movs_rows])
    movs["data"] = pd.to_datetime(movs["data"])
    movs["quantidade"] = pd.to_numeric(movs["quantidade"])
    movs["preco_unitario"] = pd.to_numeric(movs["preco_unitario"])

    cot_df = (
        pd.DataFrame([dict(r) for r in cot_rows])
        if cot_rows
        else pd.DataFrame(columns=["ticker", "mes", "preco_cota"])
    )
    im_df = (
        pd.DataFrame([dict(r) for r in im_rows])
        if im_rows
        else pd.DataFrame(columns=["ticker", "mes", "dy_mes", "vpa"])
    )

    desde_period = pd.Period(desde, freq="M") if desde else None
    hoje = pd.Timestamp.today()
    # Meses ja fechados (exclui mes corrente, cujo DY ainda nao esta completo)
    ultimo_mes = (hoje.to_period("M") - 1)

    # Fatores de grupamento confirmados para correcao historica
    from src.portfolio.grupamentos import get_fatores_por_ticker
    fatores = get_fatores_por_ticker(tickers_list)

    registros: list[dict] = []

    for t in tickers_list:
        t_movs = movs[movs["ticker"] == t].sort_values("data")
        if t_movs.empty:
            continue

        primeiro_mes = t_movs["data"].min().to_period("M")
        inicio = max(primeiro_mes, desde_period) if desde_period else primeiro_mes

        if inicio > ultimo_mes:
            continue

        # Reconstroi cotas e preco_medio evento a evento
        cotas_acc: int = 0
        pm_acc: float = 0.0
        eventos: list[tuple] = []  # (timestamp, cotas, preco_medio)

        for _, mv in t_movs.iterrows():
            q = int(mv["quantidade"])
            p = float(mv["preco_unitario"]) if pd.notna(mv["preco_unitario"]) else 0.0
            if mv["tipo"] == "compra":
                total = cotas_acc + q
                pm_acc = (cotas_acc * pm_acc + q * p) / total if total > 0 else p
                cotas_acc = total
            else:
                cotas_acc = max(0, cotas_acc - q)
            eventos.append((mv["data"], cotas_acc, pm_acc))

        splits = fatores.get(t, [])  # [(YYYY-MM, fator, tipo), ...] ordem cronologica

        for mes in pd.period_range(inicio, ultimo_mes, freq="M"):
            fim = mes.to_timestamp("M")
            cotas_mes, pm_mes = 0, 0.0
            for ev_data, ev_cotas, ev_pm in eventos:
                if ev_data <= fim:
                    cotas_mes, pm_mes = ev_cotas, ev_pm
                else:
                    break

            if cotas_mes <= 0:
                continue

            # Corrige cotas e preco_medio para refletir eventos de split:
            #   grupamento (reverse split N:1): N cotas antigas -> 1 nova
            #     meses ANTES do evento: multiplica cotas por N, divide pm por N
            #   desdobramento (forward split 1:N): 1 cota antiga -> N novas
            #     meses APOS o evento: multiplica cotas por N, divide pm por N
            # Em ambos os casos custo_total = cotas_corrigidas * pm_corrigido permanece igual.
            cotas_corrigidas = float(cotas_mes)
            pm_corrigido = float(pm_mes)
            for split_mes, split_fator, split_tipo in splits:
                if split_tipo == "grupamento" and str(mes) < split_mes:
                    cotas_corrigidas *= split_fator
                    pm_corrigido /= split_fator
                elif split_tipo == "desdobramento" and str(mes) >= split_mes:
                    cotas_corrigidas *= split_fator
                    pm_corrigido /= split_fator

            registros.append({
                "mes":         str(mes),
                "ticker":      t,
                "cotas":       cotas_corrigidas,
                "preco_medio": round(pm_corrigido, 4),
            })

    if not registros:
        return pd.DataFrame()

    df = pd.DataFrame(registros)

    # Junta cotacoes e inf_mensal
    if not cot_df.empty:
        df = df.merge(cot_df, on=["ticker", "mes"], how="left")
    else:
        df["preco_cota"] = None

    if not im_df.empty:
        df = df.merge(im_df, on=["ticker", "mes"], how="left")
    else:
        df["dy_mes"] = None
        df["vpa"] = None

    # Usa VPA como fallback quando nao ha cotacao de mercado
    preco_ref = df["preco_cota"].fillna(df.get("vpa"))

    df["custo_total"]         = (df["cotas"] * df["preco_medio"]).round(2)
    df["dividendo_cota"]      = (df["dy_mes"] * preco_ref).round(4)
    df["dividendo_recebido"]  = (df["cotas"] * df["dividendo_cota"]).round(2)
    df["yoc_mes"]             = (df["dividendo_recebido"] / df["custo_total"] * 100).round(4)

    cols = ["mes", "ticker", "cotas", "preco_medio", "custo_total",
            "preco_cota", "dy_mes", "dividendo_cota", "dividendo_recebido", "yoc_mes"]
    return df[[c for c in cols if c in df.columns]].sort_values(["ticker", "mes"]).reset_index(drop=True)


def export_template(path: str) -> None:
    """
    Salva um arquivo Excel com o template de importacao de carteira.

    Colunas: ticker | tipo | data | cotas | preco
    Inclui linhas de exemplo e validacao de lista suspensa para 'tipo'.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "carteira"

    # Cabecalho
    headers = ["ticker", "tipo", "data", "cotas", "preco"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Linhas de exemplo
    exemplos = [
        ["HGLG11", "compra", "2024-06-15", 100, 165.50],
        ["MXRF11", "compra", "2024-08-10", 500,   9.85],
        ["XPLG11", "compra", "2024-09-20", 200, 103.20],
        ["MXRF11", "venda",  "2026-04-01", 100,  10.20],
    ]
    example_fill = PatternFill("solid", fgColor="D9E1F2")
    for row_idx, row_data in enumerate(exemplos, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            if col_idx == 5:
                cell.number_format = "#,##0.00"

    # Validacao de lista suspensa para coluna 'tipo' (B)
    dv = DataValidation(type="list", formula1='"compra,venda"', allow_blank=False)
    dv.sqref = "B2:B10000"
    ws.add_data_validation(dv)

    # Largura das colunas
    ws.column_dimensions["A"].width = 12  # ticker
    ws.column_dimensions["B"].width = 10  # tipo
    ws.column_dimensions["C"].width = 14  # data
    ws.column_dimensions["D"].width = 8   # cotas
    ws.column_dimensions["E"].width = 12  # preco

    # Aba de instrucoes
    wi = wb.create_sheet("instrucoes")
    instrucoes = [
        ("Campo",  "Formato",        "Exemplo",     "Obrigatorio"),
        ("ticker", "Texto maiusculo", "HGLG11",      "Sim"),
        ("tipo",   "compra ou venda", "compra",      "Sim"),
        ("data",   "YYYY-MM-DD",      "2024-06-15",  "Sim"),
        ("cotas",  "Numero inteiro",  "100",         "Sim"),
        ("preco",  "Preco por cota",  "165.50",      "Sim"),
    ]
    hf = Font(bold=True)
    for r, row in enumerate(instrucoes, start=1):
        for c, val in enumerate(row, start=1):
            cell = wi.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = hf
    wi.column_dimensions["A"].width = 10
    wi.column_dimensions["B"].width = 20
    wi.column_dimensions["C"].width = 16
    wi.column_dimensions["D"].width = 12

    wb.save(path)


def import_from_excel(path: str, dry_run: bool = False) -> tuple[int, list[str]]:
    """
    Le o arquivo Excel (aba 'carteira') e registra as operacoes.

    Retorna:
        (n_sucesso, lista_de_erros)

    Erros nao interrompem o processamento: linhas validas sao inseridas
    mesmo que outras contenham erros.
    """
    df = pd.read_excel(path, sheet_name="carteira", dtype=str)
    df.columns = [c.strip().lower() for c in df.columns]

    required = {"ticker", "tipo", "data", "cotas", "preco"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Colunas ausentes no arquivo: {', '.join(sorted(missing))}")

    erros: list[str] = []
    n_sucesso = 0

    for i, row in df.iterrows():
        linha = i + 2  # +2 porque cabecalho e index 0-based

        # Ignora linhas completamente vazias
        if all(pd.isna(row[c]) or str(row[c]).strip() == "" for c in required):
            continue

        # Valida campos obrigatorios
        valores: dict = {}
        ok = True
        for campo in required:
            v = str(row.get(campo, "")).strip()
            if not v or v.lower() == "nan":
                erros.append(f"Linha {linha}: campo '{campo}' vazio")
                ok = False
        if not ok:
            continue

        ticker = str(row["ticker"]).strip().upper()
        tipo   = str(row["tipo"]).strip().lower()
        data   = str(row["data"]).strip()

        if tipo not in ("compra", "venda"):
            erros.append(f"Linha {linha} ({ticker}): tipo '{tipo}' invalido — use 'compra' ou 'venda'")
            continue

        try:
            cotas = int(float(str(row["cotas"]).strip()))
            if cotas <= 0:
                raise ValueError
        except (ValueError, TypeError):
            erros.append(f"Linha {linha} ({ticker}): 'cotas' deve ser inteiro positivo")
            continue

        try:
            preco = float(str(row["preco"]).strip().replace(",", "."))
            if preco <= 0:
                raise ValueError
        except (ValueError, TypeError):
            erros.append(f"Linha {linha} ({ticker}): 'preco' deve ser numero positivo")
            continue

        # Valida formato da data
        try:
            pd.Timestamp(data)
            # Normaliza para YYYY-MM-DD independente do formato de entrada
            data = pd.Timestamp(data).strftime("%Y-%m-%d")
        except Exception:
            erros.append(f"Linha {linha} ({ticker}): data '{data}' invalida — use YYYY-MM-DD")
            continue

        if dry_run:
            n_sucesso += 1
            continue

        try:
            if tipo == "compra":
                add_compra(ticker, cotas, preco, data)
            else:
                add_venda(ticker, cotas, preco, data)
            n_sucesso += 1
        except Exception as e:
            erros.append(f"Linha {linha} ({ticker}): {e}")

    return n_sucesso, erros


def get_movimentacoes(ticker: str | None = None) -> pd.DataFrame:
    """Retorna historico de movimentacoes, opcionalmente filtrado por ticker."""
    with connect() as conn:
        if ticker:
            rows = conn.execute(
                "SELECT * FROM movimentacoes WHERE ticker = ? ORDER BY data DESC, id DESC",
                (ticker.upper(),),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM movimentacoes ORDER BY data DESC, id DESC"
            ).fetchall()

    if not rows:
        return pd.DataFrame()
    return pd.DataFrame([dict(r) for r in rows])
